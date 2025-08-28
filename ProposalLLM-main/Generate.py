import openpyxl
import docx
import openai
import os
import requests
import json
import re  # Importing regular expressions for cleaning textexitt
from docx.shared import Inches
import shutil
from docx import Document
import io
import os
from docx.shared import Cm
import random
import time
from TenderFileProcessor import TenderFileProcessor

#
#这个具有读取ExcelB列，如果有内容则标题二+1，没有内容标题三+1，自动根据C列生成D列的应答，对应章节号自动写入F列，将对应标准需求写入到Word里，不改变文字和图片
#原始的word 1- XXX ，2- XXX
#如果标题名称不同，需要手工设置last_heading_2，last_heading_3，具体参见 def get_last_section_numbers(doc)，用于填写Excel内容


#--------------------------------需要调整的全局变量--------------------------------------------------



# 设置OpenAI API Key
openai.api_key = "自己的Open AIKey"
# 设置通义千问 API Key
API_KEY = "sk-fe0485c281964259b404907d483d3777"

# 图片生成相关配置
GENERATE_IMAGES = False  # 是否启用智能图片生成
IMAGE_CACHE = {}  # 图片缓存，避免重复生成

# 设置Word输出最大宽度为14厘米
MAX_WIDTH_CM = 14.0

# 优化后的Prompt，更加简洁专业
Prompt_Answer = """作为专业的标书撰写专家，请针对以下需求给出简洁、专业的陈述。格式要求：
1. 开头直接陈述技术能力
2. 内容控制在100字以内
3. 突出技术优势和产品特色
4. 语言正式、专业
5. 避免重复和冗余

需求："""

Prompt_Content = """作为大数据平台产品专家，请针对以下需求给出500字以内的产品功能介绍。要求：
1. 直接描述产品功能，不要开头和总结
2. 使用项目符号格式
3. 突出技术优势和实际应用场景
4. 语言专业、简洁
5. 避免重复内容

需求："""

# 针对产品同样的内容重写标书内容，以应对不同版本标书的Prompt
Prompt_RewriteContent = """请将以下产品介绍内容改写为标书格式，要求：
1. 保持技术准确性
2. 突出产品优势
3. 语言更加正式专业
4. 控制在500字以内

原文："""

# 针对不同需求，缩写需求内容，变为每个章节的小标题
Prompt_Title = """请将以下需求简化为10字以内的标题，要求：
1. 去掉标点符号
2. 突出核心功能
3. 语言简洁明了
4. 直接返回结果，不要解释

需求："""

# 优化图片描述生成
Prompt_Image = """请为以下内容生成专业的图片描述，要求：
1. 描述图片应该展示的核心元素
2. 突出技术架构或功能特点
3. 语言简洁专业
4. 控制在30字以内
5. 避免重复描述

内容："""

# 新增：HTML图片标签生成
Prompt_HTML_Image = """请根据以下内容生成一个用HTML代码绘制的专业图片，要求：
1. 根据图片描述和上下文内容，生成真正符合描述的图片
2. 如果是架构图，要绘制真实的系统架构，包含数据流、组件关系等
3. 如果是流程图，要绘制真实的业务流程，包含决策点、处理步骤等
4. 如果是数据流向图，要绘制真实的数据流向，包含数据源、处理节点、输出等
5. 使用SVG或Canvas绘制，确保图片内容专业、准确
6. 图片尺寸为600x400像素
7. 使用合适的颜色、字体、图标和布局
8. 确保图片内容与描述完全匹配

上下文内容：{context}
图片描述：{image_description}

请生成完整的HTML图片代码："""

# 标准标书目录结构
STANDARD_TOC_STRUCTURE = {
    "1": "项目概述",
    "2": "技术方案",
    "2.1": "系统架构设计",
    "2.2": "功能模块设计", 
    "2.3": "技术实现方案",
    "2.4": "系统集成方案",
    "3": "项目实施",
    "3.1": "实施计划",
    "3.2": "质量保证",
    "3.3": "风险控制",
    "4": "技术支持",
    "4.1": "技术团队",
    "4.2": "服务承诺",
    "5": "附录"
}

MoreSection = 1  # 当为1时，启用新的生成标题策略，可以直接生成3级标题并读取 def get_last_section_numbers(doc)
ReGenerateText = 0  # 当为1时，会对原始产品文档当中的文字内容进行重写
DDDAnswer = 1 # 当为1时，会生成点对点应答
key_flag = 0 # 当为1时，★ ▲ 会自动带到标题和需求描述当中
level1 = 'heading 1'  
level2 = 'heading 2'
last_heading_1 = 2 #技术标书起始的段落，Word当中应该有第一章，概述，那么技术从第二大章开始，此处就是2
last_heading_2 = 0
last_heading_3 = 0

# 设置输出文本内容为空
content_between_headings = []
#-----------------------------------------------------------------------------------------------------------------------

def clean_markdown_format(text):
    """清理文本中的Markdown格式，返回纯文本"""
    if not text:
        return text
    
    # 移除Markdown格式
    cleaned = text
    
    # 移除加粗格式 **text** -> text (处理嵌套情况)
    while re.search(r'\*\*(.*?)\*\*', cleaned):
        cleaned = re.sub(r'\*\*(.*?)\*\*', r'\1', cleaned)
    
    # 移除斜体格式 *text* -> text (处理嵌套情况)
    while re.search(r'\*([^*]+)\*', cleaned):
        cleaned = re.sub(r'\*([^*]+)\*', r'\1', cleaned)
    
    # 移除标题格式 # title -> title
    cleaned = re.sub(r'^#+\s*', '', cleaned, flags=re.MULTILINE)
    
    # 移除链接格式 [text](url) -> text
    cleaned = re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', cleaned)
    
    # 移除代码格式 `code` -> code
    cleaned = re.sub(r'`([^`]+)`', r'\1', cleaned)
    
    # 移除列表格式 - item -> item
    cleaned = re.sub(r'^[-*]\s*', '', cleaned, flags=re.MULTILINE)
    
    # 移除数字列表格式 1. item -> item
    cleaned = re.sub(r'^\d+\.\s*', '', cleaned, flags=re.MULTILINE)
    
    # 移除其他可能的Markdown格式
    cleaned = re.sub(r'~~(.*?)~~', r'\1', cleaned)  # 删除线
    cleaned = re.sub(r'^>\s*', '', cleaned, flags=re.MULTILINE)  # 引用
    cleaned = re.sub(r'^\|\s*', '', cleaned, flags=re.MULTILINE)  # 表格
    cleaned = re.sub(r'^\s*\|', '', cleaned, flags=re.MULTILINE)  # 表格
    
    # 清理多余的空行
    cleaned = re.sub(r'\n\s*\n\s*\n', '\n\n', cleaned)
    
    # 清理行首行尾的空白
    cleaned = re.sub(r'^\s+', '', cleaned, flags=re.MULTILINE)
    cleaned = re.sub(r'\s+$', '', cleaned, flags=re.MULTILINE)
    
    return cleaned.strip()

# 通义千问API调用
def call_qwen_api(prompt):
    """
    调用通义千问API
    :param prompt: 提示词
    :return: API响应结果
    """
    url = "https://dashscope.aliyuncs.com/api/v1/services/aigc/text-generation/generation"
    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json"
    }
    payload = {
        "model": "qwen-turbo",
        "input": {
            "messages": [
                {
                    "role": "user",
                    "content": prompt
                }
            ]
        }
    }
    
    try:
        response = requests.post(url, headers=headers, json=payload)
        response_data = response.json()
        
        # 添加调试信息
        print(f"API响应状态码: {response.status_code}")
        print(f"API响应内容: {response_data}")
        
        # 检查响应结构
        if response.status_code != 200:
            print(f"API调用失败: {response_data}")
            return {"output": {"text": "API调用失败，请检查网络连接和API密钥"}}
        
        # 检查响应结构
        if 'output' in response_data and 'text' in response_data['output']:
            return response_data
        elif 'choices' in response_data and len(response_data['choices']) > 0:
            # 兼容其他API格式
            return {"output": {"text": response_data['choices'][0]['message']['content']}}
        else:
            print(f"API响应格式异常: {response_data}")
            return {"output": {"text": "API响应格式异常"}}
            
    except Exception as e:
        print(f"API调用异常: {e}")
        return {"output": {"text": f"API调用异常: {e}"}}

def generate_toc_structure():
    """基于招标文件内容生成动态标书目录结构"""
    print("正在基于招标文件生成动态标书目录结构...")
    
    try:
        # 读取招标文件
        tender_file = "招标文件示例.json"
        if os.path.exists(tender_file):
            with open(tender_file, 'r', encoding='utf-8') as f:
                tender_data = json.load(f)
            
            # 基于招标文件内容生成大纲
            prompt = f"""作为标书撰写专家，请根据以下招标文件内容，生成一个完整的标书目录结构。
要求：
1. 包含项目概述、技术方案、项目实施、技术支持等主要章节
2. 技术方案部分要包含系统架构、功能模块、技术实现、系统集成等子章节
3. 项目实施部分要包含实施计划、质量保证、风险控制等子章节
4. 技术支持部分要包含技术团队、服务承诺等子章节
5. 返回格式为JSON，如：{{"1": "项目概述", "2": "技术方案", "2.1": "系统架构设计"}}
6. 章节编号要符合标准格式（1、2、2.1、2.2、3、3.1等）

招标文件内容：{json.dumps(tender_data, ensure_ascii=False, indent=2)}"""
            
            response = call_qwen_api(prompt)
            content = response['output']['text'].strip()
            
            # 尝试解析JSON响应
            try:
                # 提取JSON部分
                json_match = re.search(r'\{.*\}', content, re.DOTALL)
                if json_match:
                    dynamic_outline = json.loads(json_match.group())
                    print("成功生成动态大纲")
                    return dynamic_outline
                else:
                    print("无法解析JSON响应，使用默认大纲")
                    return STANDARD_TOC_STRUCTURE
            except json.JSONDecodeError:
                print("JSON解析失败，使用默认大纲")
                return STANDARD_TOC_STRUCTURE
        else:
            print("未找到招标文件，使用默认大纲")
            return STANDARD_TOC_STRUCTURE
    except Exception as e:
        print(f"生成动态大纲时出错: {e}")
        return STANDARD_TOC_STRUCTURE

def generate_chapter_content(chapter_id, chapter_title, requirements):
    """生成单个章节内容"""
    print(f"正在生成章节 {chapter_id}: {chapter_title}")
    
    # 直接让API根据章节标题智能判断内容类型
    content = generate_smart_content(chapter_id, chapter_title, requirements)
    
    # 保存章节文件 - 只保存txt格式
    filename = f"generated_proposal/章节_{chapter_id}_{chapter_title}.txt"
    
    # 确保目录存在
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    
    # 写入txt文件
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(f"{chapter_id}. {chapter_title}\n\n{content}")
    
    # 生成HTML版本
    html_filename = filename.replace('.txt', '.html')
    # 将内容转换为HTML格式
    content_html = content.replace('\n', '<br>')
    html_content = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{chapter_id}. {chapter_title}</title>
    <style>
        body {{ font-family: 'Microsoft YaHei', Arial, sans-serif; line-height: 1.6; margin: 40px; }}
        h1 {{ color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }}
        .proposal-image {{ max-width: 100%; height: auto; border: 1px solid #ddd; border-radius: 5px; margin: 20px 0; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        p {{ margin-bottom: 15px; text-align: justify; }}
    </style>
</head>
<body>
    <h1>{chapter_id}. {chapter_title}</h1>
    <div class="content">
        {content_html}
    </div>
</body>
</html>"""
    
    with open(html_filename, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    # 返回实际内容而不是文件名
    return content

def generate_smart_content(chapter_id, chapter_title, requirements):
    """智能生成章节内容，让API根据标题自动判断内容类型"""
    prompt = f"""作为标书撰写专家，请根据招标文件的具体需求生成章节内容。

章节信息：
- 章节编号：{chapter_id}
- 章节标题：{chapter_title}

招标文件需求内容：
{requirements}

重要要求：
1. 必须基于上述招标文件的具体需求来生成内容，不要使用通用的技术描述
2. 内容要针对广西智慧海洋监管服务平台建设项目的具体需求
3. 如果招标文件中有相关技术需求，要直接引用和回应这些需求
4. 避免使用通用的Spring Cloud、Redis、ELK等描述，要针对海洋监管平台的具体技术需求
5. 内容控制在800字以内
6. 在适当位置添加1-2个图片描述，格式为"（此处需[详细图片描述]的图片）"
7. 不要使用任何Markdown格式，只输出纯文本
8. 语言正式专业，突出技术优势

请根据章节标题"{chapter_title}"和招标文件需求生成相关内容："""
    
    try:
        response = call_qwen_api(prompt)
        content = response['output']['text'].strip()
        
        # 检查API返回的内容是否有效
        if not content or content.startswith("API调用失败") or content.startswith("API响应格式异常"):
            raise Exception("API返回无效内容")
        
        # 移除重复的标题
        content = re.sub(rf'^.*?{re.escape(chapter_title)}.*?\n', '', content, flags=re.MULTILINE)
        # 清理Markdown格式
        content = clean_markdown_format(content)
        
        # 将图片描述替换为HTML img标签
        content = replace_image_descriptions_with_html(content)
        
        return content
    except Exception as e:
        print(f"生成章节内容时出错: {e}")
        raise e

def generate_fallback_content(chapter_id, chapter_title, requirements):
    """生成fallback内容"""
    fallback_templates = {
        "项目概述": f"""本项目旨在建设一个现代化的智慧海洋监管服务平台，为海洋监管部门提供全面的技术支撑。

项目背景：
随着海洋经济的快速发展，传统的监管方式已无法满足现代海洋管理的需求。本项目将运用先进的信息技术，构建智能化、数字化的海洋监管体系。

建设目标：
1. 建立统一的海洋监管数据平台
2. 实现海洋监管业务的智能化处理
3. 提升监管效率和决策支持能力
4. 保障海洋生态环境安全

（此处需海洋监管平台整体架构图：展示平台的整体技术架构，包含数据采集、处理、分析、展示等各个层次的技术组件）""",
        
        "技术方案": f"""本项目采用先进的技术架构，确保系统的稳定性、可扩展性和安全性。

技术架构：
1. 采用微服务架构，实现模块化设计
2. 使用容器化技术，提高部署效率
3. 建立分布式数据处理能力
4. 实现高可用性和容灾备份

核心技术：
- 大数据处理技术
- 人工智能算法
- 地理信息系统
- 实时数据处理

（此处需技术架构详细图：展示系统的技术架构，包含前端、后端、数据库、中间件等各个技术组件的关系和数据流向）""",
        
        "项目实施": f"""项目实施将严格按照项目管理标准执行，确保项目按时、按质、按预算完成。

实施计划：
第一阶段（1-2个月）：需求调研和系统设计
第二阶段（3-6个月）：核心功能开发
第三阶段（7-8个月）：系统集成和测试
第四阶段（9-10个月）：部署上线和培训

质量保证：
1. 建立完善的质量管理体系
2. 实施代码审查和测试
3. 定期进行项目进度评估
4. 建立风险控制机制

（此处需项目实施流程图：展示项目的实施阶段、时间节点、关键里程碑和交付物）""",
        
        "技术支持": f"""我们拥有专业的技术团队和完善的服务体系，为客户提供全方位的技术支持。

技术团队：
- 项目经理：负责项目整体协调
- 架构师：负责系统架构设计
- 开发工程师：负责功能开发
- 测试工程师：负责质量保证
- 运维工程师：负责系统运维

服务承诺：
1. 7×24小时技术支持
2. 快速响应客户需求
3. 定期系统维护和升级
4. 提供技术培训服务

（此处需技术团队组织结构图：展示技术团队的组成结构、职责分工和协作关系）"""
    }
    
    # 根据章节标题选择合适的模板
    for key, template in fallback_templates.items():
        if key in chapter_title:
            return template
    
    # 默认模板
    return f"""我们将为{chapter_title}提供专业的技术解决方案，确保项目成功实施。

技术特点：
1. 采用先进的技术架构
2. 确保系统的稳定性和可靠性
3. 提供完善的技术支持
4. 满足项目的各项需求

实施保障：
- 专业的技术团队
- 完善的质量管理体系
- 及时的技术支持服务
- 持续的优化升级

（此处需{chapter_title}技术实现图：展示该章节涉及的核心技术实现方案，包含技术架构、数据流向、关键组件等详细技术实现图）"""

def generate_html_image_tag(context, image_description):
    """
    调用通义千问API生成用HTML代码绘制的专业图片
    :param context: 上下文内容
    :param image_description: 图片描述
    :return: HTML图片代码
    """
    prompt = Prompt_HTML_Image.format(
        context=context,
        image_description=image_description
    )
    
    try:
        response = call_qwen_api(prompt)
        html_code = response['output']['text'].strip()
        
        # 确保返回的是有效的HTML代码
        if not (html_code.startswith('<svg') or html_code.startswith('<canvas') or html_code.startswith('<div')):
            raise Exception("API返回的内容不是有效的HTML代码")
        
        return html_code
    except Exception as e:
        print(f"生成HTML图片代码时出错: {e}")
        raise e



def replace_image_descriptions_with_html(content):
    """
    将内容中的图片描述替换为HTML img标签
    :param content: 原始内容
    :return: 替换后的内容
    """
    # 匹配图片描述的正则表达式
    pattern = r'（此处需\[(.*?)\]的图片）'
    
    def replace_match(match):
        image_description = match.group(1)
        # 获取上下文（前后100个字符）
        start = max(0, match.start() - 100)
        end = min(len(content), match.end() + 100)
        context = content[start:end]
        
        # 生成HTML img标签
        html_tag = generate_html_image_tag(context, image_description)
        return html_tag
    
    # 替换所有匹配的图片描述
    updated_content = re.sub(pattern, replace_match, content)
    return updated_content


def merge_chapters_to_final_document():
    """将所有章节合并为最终标书"""
    print("正在合并章节生成最终标书...")
    
    # 生成txt版本的最终标书
    txt_content = ["技术标书", "", "目录", ""]
    for section_id, section_title in STANDARD_TOC_STRUCTURE.items():
        if "." in section_id:
            txt_content.append(f"    {section_id}. {section_title}")
        else:
            txt_content.append(f"  {section_id}. {section_title}")
    
    txt_content.extend(["", "=" * 50, ""])
    
    # 合并所有章节的txt内容
    chapter_files = []
    for section_id, section_title in STANDARD_TOC_STRUCTURE.items():
        filename = f"generated_proposal/章节_{section_id}_{section_title}.txt"
        if os.path.exists(filename):
            chapter_files.append((section_id, filename))
    
    # 按章节顺序合并
    chapter_files.sort(key=lambda x: x[0])
    
    for section_id, filename in chapter_files:
        with open(filename, 'r', encoding='utf-8') as f:
            txt_content.append(f.read())
            txt_content.append("")
    
    # 确保目录存在
    os.makedirs("generated_proposal", exist_ok=True)
    
    with open("generated_proposal/最终标书.txt", 'w', encoding='utf-8') as f:
        f.write('\n'.join(txt_content))
    
    # 生成HTML版本的最终标书
    html_content = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>技术标书</title>
    <style>
        body { font-family: 'Microsoft YaHei', Arial, sans-serif; line-height: 1.6; margin: 40px; }
        h1 { color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }
        h2 { color: #34495e; margin-top: 30px; }
        .proposal-image { max-width: 100%; height: auto; border: 1px solid #ddd; border-radius: 5px; margin: 20px 0; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
        p { margin-bottom: 15px; text-align: justify; }
        .toc { background-color: #f8f9fa; padding: 20px; border-radius: 5px; margin: 20px 0; }
        .toc h2 { margin-top: 0; }
        .toc-item { margin: 5px 0; }
        .chapter { margin: 30px 0; }
    </style>
</head>
<body>
    <h1>技术标书</h1>
    
    <div class="toc">
        <h2>目录</h2>"""
    
    # 添加目录
    for section_id, section_title in STANDARD_TOC_STRUCTURE.items():
        if "." in section_id:
            html_content += f'<div class="toc-item">    {section_id}. {section_title}</div>'
        else:
            html_content += f'<div class="toc-item">  {section_id}. {section_title}</div>'
    
    html_content += """
    </div>
    
    <hr>
    """
    
    # 合并所有章节的HTML内容
    for section_id, filename in chapter_files:
        html_filename = filename.replace('.txt', '.html')
        if os.path.exists(html_filename):
            with open(html_filename, 'r', encoding='utf-8') as f:
                chapter_html = f.read()
                # 提取章节内容部分
                content_start = chapter_html.find('<div class="content">')
                content_end = chapter_html.find('</div>', content_start)
                if content_start != -1 and content_end != -1:
                    chapter_content = chapter_html[content_start + 20:content_end]
                    html_content += f'<div class="chapter">{chapter_content}</div>'
    
    html_content += """
</body>
</html>"""
    
    with open("generated_proposal/最终标书.html", 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print("最终标书已生成：generated_proposal/最终标书.txt 和 generated_proposal/最终标书.html")

def main():
    """主函数 - 分阶段生成标书"""
    print("=== 标书生成系统 ===")
    print("1. 处理招标文件")
    print("2. 生成目录结构")
    print("3. 分章节生成内容")
    print("4. 合并为最终标书")
    print("==================")
    
    # 阶段1：处理招标文件
    print("开始处理招标文件...")
    processor = TenderFileProcessor()
    
    # 查找招标文件
    tender_files = []
    # 搜索当前目录和子目录
    for root, dirs, files in os.walk('.'):
        for file in files:
            if file.endswith(('.docx', '.xlsx', '.xls', '.json', '.doc')) and ('招标' in file or 'tender' in file.lower()):
                tender_files.append(os.path.join(root, file))
    
    print(f"找到的招标文件: {tender_files}")
    
    if not tender_files:
        print("未找到招标文件，使用默认需求对应表...")
        # 使用原有的需求对应表
        excel_file = "需求对应表.xlsx"
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        
        requirements = []
        for row in sheet.iter_rows(min_row=2):
            c_column_content = row[2].value
            if c_column_content:
                requirements.append(c_column_content)
        
        requirements_text = "\n".join(requirements)
    else:
        print(f"找到招标文件: {tender_files}")
        tender_file = tender_files[0]
        requirements_data = processor.process_tender_file(tender_file)
        
        if not requirements_data:
            print("无法解析招标文件，使用默认需求对应表...")
            # 使用原有的需求对应表
            excel_file = "需求对应表.xlsx"
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            
            requirements = []
            for row in sheet.iter_rows(min_row=2):
                c_column_content = row[2].value
                if c_column_content:
                    requirements.append(c_column_content)
            
            requirements_text = "\n".join(requirements)
        else:
            # 转换需求数据格式
            requirements = []
            for req in requirements_data:
                requirements.append(req['content'])
            
            requirements_text = "\n".join(requirements)
            
            # 生成需求摘要
            summary = processor.generate_requirements_summary(requirements_data)
            print(f"提取需求统计:")
            print(f"总需求数: {summary['total_requirements']}")
            print(f"技术需求: {summary['technical_requirements']}")
            print(f"业务需求: {summary['business_requirements']}")
            print(f"功能需求: {summary['functional_requirements']}")
            print(f"高优先级: {summary['high_priority']}")
    
    # 阶段2：生成目录
    toc_structure = generate_toc_structure()
    
    # 阶段3：分章节生成
    print("开始生成各章节内容...")
    
    # 生成各章节
    for section_id, section_title in toc_structure.items():
        generate_chapter_content(section_id, section_title, requirements_text)
    
    # 阶段4：合并为最终标书
    merge_chapters_to_final_document()
    
    print("标书生成完成！")

if __name__ == "__main__":
    main()